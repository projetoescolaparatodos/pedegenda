import datetime as dt

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _parse_datahora(value: str) -> tuple[bool, dt.date | None, dt.time | None]:
    """Converte "DD/MM/AAAA - ..." em componentes ordenáveis.

    Retorna (tem_data, data, hora). Itens "A Definir" ficam com tem_data=False.
    """

    value = (value or "").strip()
    if not value or value.lower().startswith("a definir"):
        return (False, None, None)

    # Esperado: "13/01/2026 - Manhã" ou "16/01/2026 - 07:30"
    if " - " not in value:
        return (False, None, None)

    date_part, period_part = value.split(" - ", 1)
    try:
        parsed_date = dt.datetime.strptime(date_part.strip(), "%d/%m/%Y").date()
    except ValueError:
        return (False, None, None)

    period_part = period_part.strip().lower()
    if ":" in period_part:
        try:
            parsed_time = dt.datetime.strptime(period_part, "%H:%M").time()
        except ValueError:
            parsed_time = None
    else:
        # Mapeamento simples de períodos para horários aproximados
        period_to_time = {
            "manhã": dt.time(9, 0),
            "durante o dia": dt.time(13, 0),
            "fim do dia": dt.time(18, 0),
        }
        parsed_time = period_to_time.get(period_part)

    return (True, parsed_date, parsed_time)


def _sort_key(row: list[str], original_index: int) -> tuple[int, dt.date, dt.time, int]:
    has_date, parsed_date, parsed_time = _parse_datahora(row[1])
    if not has_date or parsed_date is None:
        # "A Definir" vai pro final, mantendo ordem relativa
        return (1, dt.date.max, dt.time.max, original_index)
    return (0, parsed_date, parsed_time or dt.time(12, 0), original_index)


def main() -> None:
    # Criar workbook e sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agenda Visitas"

    # Cabeçalhos
    headers = [
        "Local",
        "Data/Hora",
        "Tarefas/Checklist",
        "Prioridade",
        "Responsável",
        "Status",
        "Notas",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    ws.row_dimensions[1].height = 24

    # Dados da agenda
    # Formato: [Local, Data/Hora, Tarefas/Checklist, Prioridade, Responsável, Status, Notas]
    data = [
        [
            "Róbson",
            "13/01/2026 - Manhã",
            "Cadastrar produtos\nVincular conta",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Bom Gosto",
            "13/01/2026 - Durante o dia",
            "Mostrar o catálogo já cadastrado\nVerificar valores\nVincular conta",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Amazonina",
            "13/01/2026 - Manhã",
            "Passar e verificar/atualizar",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Açaí do Vizinho da Esquina",
            "13/01/2026 - Durante o dia",
            "Cadastrar produtos\nVincular conta",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Granja",
            "15/01/2026 - Durante o dia",
            "Apresentar o sistema\nCadastrar os produtos",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Açougue do Marquinhos",
            "15/01/2026 - Fim do dia",
            "Apresentar a plataforma\nCadastrar produtos (se possível)",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Adelson dos Salgados",
            "16/01/2026 - 07:30",
            "Cadastrar produtos\nAlinhar com as novidades\nVincular conta do restaurante",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Padeiro (pra roça)",
            "A Definir - Outro dia da semana",
            "Visitar para marcar reunião",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Teresinha",
            "A Definir - Marcar dia",
            "Atualizar preços",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Valéria do Nunes",
            "A Definir",
            "Vincular conta do restaurante\nAlinhar com atualizações do sistema",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Taila da Drogaria Vitória",
            "A Definir",
            "Vincular conta\nAlinhar com atualizações do sistema",
            "",
            "",
            "Pendente",
            "",
        ],
        [
            "Açougue 1 (nome esquecido)",
            "A Definir",
            "Vincular conta\nAlinhar com atualizações do sistema",
            "",
            "",
            "Pendente",
            "",
        ],
    ]

    # Ordenar por data/hora ("A Definir" vai pro final), mantendo ordem original quando empatar
    indexed_rows = list(enumerate(data))
    indexed_rows.sort(key=lambda pair: _sort_key(pair[1], pair[0]))
    data = [row for _, row in indexed_rows]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Preencher os dados
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

        # Deixar as linhas mais altas para textos com quebra de linha
        ws.row_dimensions[row_idx].height = 60

    # Ajustar largura das colunas
    column_widths = [25, 22, 50, 12, 18, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar a primeira linha (cabeçalho)
    ws.freeze_panes = "A2"

    # Tabela com estilo (listras + filtro no cabeçalho)
    last_row = 1 + len(data)
    table = Table(displayName="AgendaVisitas", ref=f"A1:G{last_row}")
    table_style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Formatação condicional (Status)
    # Coluna F (Status): Pendente=vermelho, Em andamento=amarelo, Concluído=verde
    status_range = f"F2:F{last_row}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=["UPPER($F2)=\"PENDENTE\""],
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=["UPPER($F2)=\"EM ANDAMENTO\""],
            fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=["UPPER($F2)=\"CONCLUÍDO\""],
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        ),
    )

    # Formatação condicional (Prioridade)
    # Coluna D (Prioridade): Alta=vermelho, Média=amarelo, Baixa=verde
    prioridade_range = f"D2:D{last_row}"
    ws.conditional_formatting.add(
        prioridade_range,
        FormulaRule(
            formula=["UPPER($D2)=\"ALTA\""],
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        prioridade_range,
        FormulaRule(
            formula=["UPPER($D2)=\"MÉDIA\""],
            fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        prioridade_range,
        FormulaRule(
            formula=["UPPER($D2)=\"MEDIA\""],
            fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        prioridade_range,
        FormulaRule(
            formula=["UPPER($D2)=\"BAIXA\""],
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        ),
    )

    # Salvar o arquivo
    output_name = "Agenda_Visitas.xlsx"
    wb.save(output_name)
    print(f"Arquivo '{output_name}' criado com sucesso!")


if __name__ == "__main__":
    main()
